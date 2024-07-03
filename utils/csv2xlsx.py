#!/usr/bin/env python3

import sys
import pandas as pd
import openpyxl
import collections
from pathlib import Path
from pdkit import df_contains

def fit_text(excel_fnam):
  wb = openpyxl.load_workbook(excel_fnam)
  for ws in wb:
    for cell in ws[1]:
      ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].bestFit = True
    for col in ws.iter_cols():
      for cell in col:
        cell.alignment = openpyxl.styles.Alignment(wrap_text = True)
  wb.save(excel_fnam)

def dfs2xlsx(sheets, output, engine_kwargs = {}, to_excel_kwargs = {}):
  if not isinstance(sheets, dict):
    if isinstance(sheets, pd.DataFrame):
      sheets = { 'Sheet 1': sheets }
    else:
      print(f'Variable sheets is neither dict nor DataFrame', file = sys.stderr)
      sys.exit(1)

  #re https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_excel.html
  with pd.ExcelWriter(output, **engine_kwargs) as writer:
    for name, df in sheets.items():
      assert isinstance(df, pd.DataFrame), f'Variable "df" has non-DataFrame type {type(df)} with value {df}'
      if 'na_rep' in to_excel_kwargs: assert not df_contains(df, to_excel_kwargs['na_rep'])
      df.to_excel(writer, sheet_name = name, **to_excel_kwargs)
  fit_text(output)

def csvs2xslx(csvs, output, csv_kwargs, engine_kwargs, to_excel_kwargs):
  if isinstance(csvs, str):
    csvs = [csvs]
  elif not isinstance(csvs, collections.abc.Sequence):
    print(f'csvs is neither str nor sequence', file = sys.stderr)
    sys.exit(1)
  dfs2xlsx({Path(x).stem: pd.read_csv(x, **csv_kwargs) for x in csvs}, output, engine_kwargs, to_excel_kwargs)


if __name__ == '__main__':
  import argparse
  parser = argparse.ArgumentParser()
  parser.add_argument('csvs',
                      type = str,
                      nargs = '+',
                      help = 'CSV files to convert into XLSX sheets')
  parser.add_argument('--output', '-o',
                      default = 'output.xlsx',
                      help = 'Filename for output spreadsheet')
  parser.add_argument('--freeze-row', default = 0, type = int)
  parser.add_argument('--freeze-col', default = 0, type = int)
  parser.add_argument('--na-rep', default = None)
  parser.add_argument('--index-col', default = None, type = int)
  parser.add_argument('--float-format', default = None, type = str)
  parser.add_argument('--separator', '--sep', default = None, type = str)
  args = parser.parse_args()
  engine_kwargs = {
    'mode': 'w',
    'engine': 'openpyxl',
  }
  to_excel_kwargs = {
    'index': None,
    'freeze_panes': (args.freeze_row, args.freeze_col),
  }
  csv_kwargs = {}
  if not args.na_rep is None: to_excel_kwargs['na_rep'] = args.na_rep
  if not args.float_format is None: to_excel_kwargs['float_format'] = args.float_format
  if not args.index_col is None:
    csv_kwargs['index_col'] = args.index_col
    to_excel_kwargs['index'] = True
  if not args.separator is None:
    if args.separator == 'TAB':
      csv_kwargs['sep'] = '\t'
    else:
      csv_kwargs['sep'] = args.separator
  csvs2xslx(args.csvs, args.output, csv_kwargs, engine_kwargs, to_excel_kwargs)
